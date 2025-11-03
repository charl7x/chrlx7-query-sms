"""
Excel导出模块
将短信查询结果导出为Excel文件
"""
from typing import List, Dict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel导出器"""
    
    def __init__(self):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "短信发送明细"
    
    def export(self, data: List[Dict], output_file: str):
        """
        导出数据到Excel文件
        
        Args:
            data: 短信记录列表
            output_file: 输出文件路径
        """
        if not data:
            print("没有数据可导出")
            return
        
        # 设置表头
        self._setup_headers()
        
        # 写入数据
        self._write_data(data)
        
        # 调整列宽
        self._adjust_column_widths()
        
        # 保存文件
        self.workbook.save(output_file)
        print(f"成功导出 {len(data)} 条记录到文件: {output_file}")
    
    def _setup_headers(self):
        """设置表头"""
        headers = ['手机号', '发送时间', '发送状态', '短信内容']
        
        # 表头样式
        header_font = Font(bold=True, size=12, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    def _write_data(self, data: List[Dict]):
        """
        写入数据行
        
        Args:
            data: 短信记录列表
        """
        # 数据对齐
        alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        for row_num, record in enumerate(data, 2):
            # 手机号
            cell = self.worksheet.cell(row=row_num, column=1)
            cell.value = record.get('phone_number', '')
            cell.alignment = alignment
            
            # 发送时间
            cell = self.worksheet.cell(row=row_num, column=2)
            cell.value = record.get('send_time', '')
            cell.alignment = alignment
            
            # 发送状态
            cell = self.worksheet.cell(row=row_num, column=3)
            status = record.get('status', '')
            cell.value = status
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 根据状态设置颜色
            if '成功' in status:
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            elif '失败' in status:
                cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            # 短信内容
            cell = self.worksheet.cell(row=row_num, column=4)
            cell.value = record.get('content', '')
            cell.alignment = alignment
    
    def _adjust_column_widths(self):
        """自动调整列宽"""
        column_widths = {
            1: 15,  # 手机号
            2: 20,  # 发送时间
            3: 12,  # 发送状态
            4: 50   # 短信内容
        }
        
        for col_num, width in column_widths.items():
            column_letter = get_column_letter(col_num)
            self.worksheet.column_dimensions[column_letter].width = width
        
        # 设置行高
        self.worksheet.row_dimensions[1].height = 25  # 表头行高


def export_to_excel(data: List[Dict], output_file: str):
    """
    便捷函数：导出数据到Excel
    
    Args:
        data: 短信记录列表
        output_file: 输出文件路径
    """
    exporter = ExcelExporter()
    exporter.export(data, output_file)

